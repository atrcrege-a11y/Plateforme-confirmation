"""dashboard.py — Agrégat secrétariat + export Excel (logique pure, testable)."""
import io
import os

DEFAULT_ADMIN_TOKEN = "admin-local"


def admin_token():
    return os.environ.get("ADMIN_TOKEN", DEFAULT_ADMIN_TOKEN)


def is_admin(token):
    return token == admin_token()


def agreger(conn):
    competitions = conn.execute(
        "SELECT id, nom, categorie, format, arme, genre, date, lieu, date_limite "
        "FROM competition ORDER BY date_limite, nom"
    ).fetchall()
    out = []
    for comp in competitions:
        confs = conn.execute(
            "SELECT cf.id, cf.club_id, cf.token, cf.statut, cf.date_confirmation, "
            "       cf.confirme_par_email, cf.corrige_par, cf.date_correction, "
            "       cl.nom AS club_nom, cl.email AS club_email "
            "FROM confirmation cf JOIN club cl ON cl.id = cf.club_id "
            "WHERE cf.competition_id = ? ORDER BY cl.nom",
            (comp["id"],),
        ).fetchall()
        clubs = []
        n_confirmes = n_presents = n_arbitres = 0
        for cf in confs:
            presents = conn.execute(
                "SELECT COUNT(*) AS n FROM participation_tireur "
                "WHERE confirmation_id = ? AND present = 1",
                (cf["id"],),
            ).fetchone()["n"]
            arbitres = conn.execute(
                "SELECT COUNT(*) AS n FROM arbitre WHERE confirmation_id = ?",
                (cf["id"],),
            ).fetchone()["n"]
            tireurs = [
                {
                    "nom": t["nom"], "prenom": t["prenom"], "equipe": t["equipe"],
                    "present": t["present"], "taille_veste": t["taille_veste"],
                    "categorie_age": t["categorie_age"],
                }
                for t in conn.execute(
                    "SELECT q.nom, q.prenom, q.equipe, "
                    "       pt.present, pt.taille_veste, pt.categorie_age "
                    "FROM participation_tireur pt "
                    "LEFT JOIN qualifie q ON q.id = pt.qualifie_id "
                    "WHERE pt.confirmation_id = ? "
                    "ORDER BY q.equipe, q.rang, q.nom",
                    (cf["id"],),
                ).fetchall()
            ]
            attendus = [
                {"qualifie_id": q["id"], "nom": q["nom"], "prenom": q["prenom"],
                 "section": q["section"], "equipe": q["equipe"], "rang": q["rang"]}
                for q in conn.execute(
                    "SELECT id, nom, prenom, section, equipe, rang FROM qualifie "
                    "WHERE competition_id = ? AND club_id = ? ORDER BY equipe, rang, nom",
                    (comp["id"], cf["club_id"]),
                ).fetchall()
            ]
            saisie = {
                pt["qualifie_id"]: pt for pt in conn.execute(
                    "SELECT qualifie_id, present, taille_veste, categorie_age "
                    "FROM participation_tireur WHERE confirmation_id = ?",
                    (cf["id"],),
                ).fetchall()
            }
            roster = []
            for at in attendus:
                pt = saisie.get(at["qualifie_id"])
                roster.append({**at, "saisi": pt is not None,
                               "present": (pt["present"] if pt else None),
                               "taille_veste": (pt["taille_veste"] if pt else None),
                               "categorie_age": (pt["categorie_age"] if pt else None)})
            if cf["statut"] == "confirmee":
                n_confirmes += 1
            n_presents += presents
            n_arbitres += arbitres
            clubs.append({
                "club": cf["club_nom"], "email": cf["club_email"],
                "statut": cf["statut"], "date_confirmation": cf["date_confirmation"],
                "confirme_par_email": cf["confirme_par_email"],
                "corrige_par": cf["corrige_par"], "date_correction": cf["date_correction"],
                "presents": presents, "arbitres": arbitres,
                "tireurs": tireurs, "attendus": roster,
                "attendus_total": len(attendus), "token": cf["token"],
                "confirmation_id": cf["id"],
            })
        total = len(confs)
        out.append({
            "id": comp["id"], "nom": comp["nom"], "categorie": comp["categorie"],
            "format": comp["format"], "arme": comp["arme"], "genre": comp["genre"],
            "date": comp["date"], "lieu": comp["lieu"],
            "date_limite": comp["date_limite"], "clubs_total": total,
            "clubs_confirmes": n_confirmes, "clubs_en_attente": total - n_confirmes,
            "tireurs_presents": n_presents, "arbitres": n_arbitres, "clubs": clubs,
        })
    return out


def participations_export(conn, competition_id):
    tireurs = conn.execute(
        "SELECT cl.nom AS club, q.equipe, q.section, q.nom, q.prenom, "
        "       pt.present, pt.taille_veste, pt.categorie_age "
        "FROM participation_tireur pt "
        "JOIN confirmation cf ON cf.id = pt.confirmation_id "
        "JOIN club cl ON cl.id = cf.club_id "
        "LEFT JOIN qualifie q ON q.id = pt.qualifie_id "
        "WHERE cf.competition_id = ? AND cf.statut = 'confirmee' "
        "ORDER BY cl.nom, q.equipe, q.rang",
        (competition_id,),
    ).fetchall()
    arbitres = conn.execute(
        "SELECT cl.nom AS club_confirmant, a.nom, a.prenom, a.club, a.niveau "
        "FROM arbitre a "
        "JOIN confirmation cf ON cf.id = a.confirmation_id "
        "JOIN club cl ON cl.id = cf.club_id "
        "WHERE cf.competition_id = ? AND cf.statut = 'confirmee' "
        "ORDER BY cl.nom, a.nom",
        (competition_id,),
    ).fetchall()
    return tireurs, arbitres


def construire_xlsx(comp_nom, tireurs, arbitres):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    gras = Font(bold=True)
    ws = wb.active
    ws.title = "Tireurs"
    entetes_t = ["Club", "Équipe", "Section", "Nom", "Prénom",
                 "Présent", "Veste", "Catégorie d'âge"]
    ws.append(entetes_t)
    for c in ws[1]:
        c.font = gras
    for t in tireurs:
        ws.append([
            t["club"], t["equipe"], t["section"], t["nom"], t["prenom"],
            "Oui" if t["present"] else "Non",
            t["taille_veste"] or "", t["categorie_age"] or "",
        ])
    wa = wb.create_sheet("Arbitres")
    entetes_a = ["Club confirmant", "Nom", "Prénom", "Club arbitre", "Niveau"]
    wa.append(entetes_a)
    for c in wa[1]:
        c.font = gras
    for a in arbitres:
        wa.append([a["club_confirmant"], a["nom"], a["prenom"], a["club"], a["niveau"]])
    for sheet, entetes in ((ws, entetes_t), (wa, entetes_a)):
        for i, _ in enumerate(entetes, start=1):
            sheet.column_dimensions[chr(64 + i)].width = 20
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

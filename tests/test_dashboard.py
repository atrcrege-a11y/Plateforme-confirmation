"""Tests dashboard — auth session, agrégat, détail tireurs, export xlsx."""
import io
import os
import tempfile

import pytest

import migrate
import seed_demo
import auth
from db import get_connection
import app as app_module


@pytest.fixture()
def client():
    fd, path = tempfile.mkstemp(suffix=".db")
    os.close(fd)
    os.environ["PLATEFORME_DB"] = path
    migrate.migrate(path)
    conn = get_connection(path)
    seed_demo.seed(conn)
    auth.creer_utilisateur(conn, "admin@x.fr", "Admin", "pw-admin", "admin")
    conn.close()
    app_module.app.config["TESTING"] = True
    c = app_module.app.test_client()
    c.post("/login", data={"email": "admin@x.fr", "mdp": "pw-admin"})  # admin connecté
    yield c
    os.environ.pop("PLATEFORME_DB", None)
    os.remove(path)


def test_agregat_non_cloisonne_et_progression(client):
    data = client.get("/api/dashboard").get_json()
    comp1 = next(c for c in data["competitions"] if c["id"] == 1)
    assert comp1["clubs_total"] == 2
    assert comp1["clubs_confirmes"] == 0 and comp1["clubs_en_attente"] == 2
    g = client.get("/api/c/tok-chalons").get_json()
    mine = [q["id"] for q in g["qualifies"] if q["mine"]]
    client.post("/api/confirm/tok-chalons", json={
        "participations": [{"qualifie_id": mine[0], "present": True, "taille_veste": "M"},
                           {"qualifie_id": mine[1], "present": False}]})
    data = client.get("/api/dashboard").get_json()
    comp1 = next(c for c in data["competitions"] if c["id"] == 1)
    assert comp1["clubs_confirmes"] == 1 and comp1["clubs_en_attente"] == 1
    assert comp1["tireurs_presents"] == 1


def test_detail_tireurs_par_club(client):
    g = client.get("/api/c/tok-chalons").get_json()
    mine = [q["id"] for q in g["qualifies"] if q["mine"]]
    client.post("/api/confirm/tok-chalons", json={
        "participations": [{"qualifie_id": mine[0], "present": True, "taille_veste": "M"},
                           {"qualifie_id": mine[1], "present": False}]})
    data = client.get("/api/dashboard").get_json()
    comp1 = next(c for c in data["competitions"] if c["id"] == 1)
    club = next(k for k in comp1["clubs"] if k["statut"] == "confirmee")
    assert len(club["tireurs"]) == 2
    present = next(t for t in club["tireurs"] if t["present"])
    assert present["taille_veste"] == "M" and present["nom"]
    assert any(not t["present"] for t in club["tireurs"])
    en_attente = next(k for k in comp1["clubs"] if k["statut"] != "confirmee")
    assert en_attente["tireurs"] == []


def test_export_xlsx(client):
    g = client.get("/api/c/tok-vet-chalons").get_json()
    mine = [q["id"] for q in g["qualifies"] if q["mine"]]
    client.post("/api/confirm/tok-vet-chalons", json={
        "participations": [{"qualifie_id": mine[0], "present": True, "categorie_age": "V1-V2"},
                           {"qualifie_id": mine[1], "present": False}],
        "arbitres": [{"nom": "A", "prenom": "B", "club": "C", "niveau": "national"}]})
    r = client.get("/api/dashboard/export/2")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["Content-Type"]
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.data))
    assert wb.sheetnames == ["Tireurs", "Arbitres"]
    assert wb["Tireurs"].max_row == 3
    assert wb["Arbitres"].max_row == 2
    assert wb["Arbitres"].cell(2, 5).value == "national"


def test_roster_attendus_et_arme_genre(client):
    data = client.get("/api/dashboard").get_json()
    comp1 = next(c for c in data["competitions"] if c["id"] == 1)
    # champs de regroupement exposés
    assert "arme" in comp1 and "genre" in comp1
    # avant toute confirmation : attendus présents, rien de saisi
    chalons = next(k for k in comp1["clubs"] if k["club"].startswith("C.E. de Châlons"))
    assert chalons["attendus_total"] == 2
    assert len(chalons["attendus"]) == 2
    assert all(a["saisi"] is False and a["present"] is None for a in chalons["attendus"])
    # après confirmation : le roster reflète la saisie, attendus_total inchangé
    g = client.get("/api/c/tok-chalons").get_json()
    mine = [q["id"] for q in g["qualifies"] if q["mine"]]
    client.post("/api/confirm/tok-chalons", json={
        "participations": [{"qualifie_id": mine[0], "present": True, "taille_veste": "M"},
                           {"qualifie_id": mine[1], "present": False}]})
    data = client.get("/api/dashboard").get_json()
    comp1 = next(c for c in data["competitions"] if c["id"] == 1)
    chalons = next(k for k in comp1["clubs"] if k["club"].startswith("C.E. de Châlons"))
    assert chalons["attendus_total"] == 2
    assert all(a["saisi"] for a in chalons["attendus"])
    assert sum(1 for a in chalons["attendus"] if a["present"]) == 1
